import openpyxl
import os
import csv
from datetime import date


def list_files(folder):
    r = []
    for root, dirs, files in os.walk(folder):
        for name in files:
            r.append(os.path.join(root, name))
    return r


def check_controls(workbatch):
    try:
        wb = openpyxl.load_workbook(workbatch)
        # export_ws = wb.get_sheet_by_name('Export')
        setup_ws = wb.get_sheet_by_name('Setup')

        ntcFailCount = 0
        posFailCount = 0

        # for i in range(1, 100):
        #     if str(export_ws.cell(row=i, column=1).value) == 'NTC' and \
        #         str(export_ws.cell(row=i, column=2).value) == 'COVID' and \
        #             str(export_ws.cell(row=i, column=3).value) != 'NOT DETECTED':
        #         ntcFailCount += 1
        #     if str(export_ws.cell(row=i, column=1).value) == 'POSITIVE' and \
        #             str(export_ws.cell(row=i, column=2).value) == 'COVID' and \
        #                 str(export_ws.cell(row=i, column=3).value) != 'Control Positive':
        #         posFailCount += 1

        for i in range(2, 9):
            if setup_ws.cell(row=17, column=i).value == 'NTC' and \
                (setup_ws.cell(row=21, column=i).value != 'NEGATIVE' or \
                len(setup_ws.cell(row=20, column=i).value) >=1):
                ntcFailCount += 1
            if setup_ws.cell(row=17, column=i).value == 'POSITIVE' and \
                (setup_ws.cell(row=21, column=i).value != 'Control Positive' or \
                setup_ws.cell(row=20, column=i).value <= 86.0):
                posFailCount += 1

        if ntcFailCount != 0 or posFailCount != 0:
            return workbatch, [ntcFailCount, posFailCount]
        else:
            return 'Pass'
    except:
        return 'Error'

def results(batchFails):
    timeNow = date.today()
    with open(f'LAMP control check {timeNow.strftime("%Y-%m-%d")}.csv', 'w', newline='') as f:
        w = csv.writer(f)
        w.writerows(batchFails.items())

    print(f'\n\nComplete.\n\nOutfile file generated: LAMP Control Check {timeNow.strftime("%Y-%m-%d")}.csv')


def run(folder):
    workbatches = list_files(folder)

    batchFails = {}
    batchFails['Workbatch'] = '[ntc, pos] fails'

    for workbatch in workbatches:
        print('Processing...', workbatch)
        if check_controls(workbatch) == 'Error':
            batchFails[workbatch] = 'Error'
        elif check_controls(workbatch) != 'Pass':
            batchFails[check_controls(workbatch)[0]] = check_controls(workbatch)[1]

    results(batchFails)

# run(r"C:\Users\bhsbu\dev\Work\lamp_controls\data\Workbatches")
run(input("Enter LAMP worksheet directory: "))
input("\n\nPress any key to exit...")
